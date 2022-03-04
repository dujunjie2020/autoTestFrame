import os
from openpyxl import load_workbook
from common import getPath


class readExcel():
    # 初始化操作
    def __init__(self, excelName):
        self.excelName = excelName
     #获取表格中的数据
    def getExcelCell(self):
        #获取文件路径
        excelFile = os.path.join(getPath.testCaseFile, self.excelName)
        #加载表格
        wb = load_workbook(excelFile)
        #创建空列表将获取到的数据存放到列表中
        caseList = []
        #获取表格中的表单
        for sheetName in wb.sheetnames:
            sheet = wb[sheetName]
            #获取每张表中最大行数
            maxRow = sheet.max_row
            #获取每行中需要的数据，并存放到列表中
            for i in range(2,maxRow+1):
                dictData = dict(
                    functionModule = sheet.cell(row=i,column=1).value,
                    caseID = sheet.cell(row=i,column=2).value,
                    url = sheet.cell(row=i,column=5).value,
                    headers = sheet.cell(row=i,column=7).value,
                    data = sheet.cell(row=i,column=8).value,
                    loginType = sheet.cell(row=i,column=9).value
                )
                caseList.append(dictData)
        return caseList

if __name__ == '__main__':
    # print(getPath.testCaseFile)
    caseFile = readExcel('Case_demo.xlsx')
    print(caseFile.getExcelCell())